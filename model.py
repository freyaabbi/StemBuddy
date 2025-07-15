import os
import re
import cv2
import json
import time
import pickle
import asyncio
import tempfile
import subprocess
import numpy as np
import face_recognition
import openai
import edge_tts
import speech_recognition as sr
import dlib
import smtplib
import requests
from email.mime.text import MIMEText
from imutils import face_utils
from datetime import datetime
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv

# Load .env file for secrets
load_dotenv()
openai.api_key = os.getenv("OPENAI_API_KEY")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")
WEATHER_API_KEY = os.getenv("WEATHER_API_KEY")
CITY_NAME = os.getenv("CITY_NAME", "Noida")

# Load company data
with open("datast.json", "r") as f:
    stemrobo_data = json.load(f)

# Excel setup
excel_file = "visitor_log.xlsx"
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Timestamp", "Image Filename", "To Meet", "Purpose"])
    wb.save(excel_file)

# Load face encodings
encodings_file = "encodings.pickle"
if os.path.exists(encodings_file):
    with open(encodings_file, "rb") as f:
        data = pickle.load(f)
    known_face_encodings = data["encodings"]
    known_face_names = data["names"]
else:
    known_face_encodings = []
    known_face_names = []

# Directories
IMAGE_DIR = "registered_faces"
os.makedirs(IMAGE_DIR, exist_ok=True)

def clean_text_for_tts(text):
    text = ''.join(char for char in text if char.isprintable())
    text = re.sub(r"[*_#<>[\]{}|~^]", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text

async def speak(text):
    communicate = edge_tts.Communicate(text, voice="en-IN-NeerjaNeural")
    with tempfile.NamedTemporaryFile(delete=False, suffix=".mp3") as f:
        await communicate.save(f.name)
        os.system(f"mpg123 {f.name}")

def speak_async(text):
    clean = clean_text_for_tts(text)
    asyncio.run(speak(clean))

def listen(language="en-IN", mic_index=1):
    r = sr.Recognizer()
    try:
        with sr.Microphone(device_index=mic_index) as source:
            print("[INFO] Listening...")
            r.adjust_for_ambient_noise(source, duration=3)
            audio = r.listen(source, timeout=10, phrase_time_limit=12)
    except sr.WaitTimeoutError:
        return "Unknown"
    except Exception as e:
        print(f"[ERROR] Mic issue: {e}")
        return "Unknown"
    try:
        text = r.recognize_google(audio, language=language)
        print(f"[USER SAID]: {text}")
        return text
    except:
        return "Unknown"

def capture_image(filename="face.jpg"):
    cmd = f"libcamera-still -t 5000 -o {filename} --width 640 --height 480 --nopreview"
    subprocess.run(cmd, shell=True)
    return filename if os.path.exists(filename) else None

def extract_face(image_path):
    img = cv2.imread(image_path)
    locations = face_recognition.face_locations(img)
    if not locations:
        return None, None, None
    top, right, bottom, left = locations[0]
    face_img = img[top:bottom, left:right]
    return face_img, locations[0], img

def ask_openai(prompt):
    try:
        response = openai.ChatCompletion.create(
            model="gpt-4.1",
            messages=[
                {"role": "system", "content": "You are a helpful assistant."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=100
        )
        return response["choices"][0]["message"]["content"]
    except Exception as e:
        return "Sorry, I couldn't process that."

def send_email_notification(to_name, from_name, purpose):
    sender = "freya.abbi@tevatrontech.com"
    receiver = f"{to_name.lower().replace(' ', '')}@stemrobo.com"
    subject = f"Visitor Alert: {from_name} wants to meet you"
    body = f"{from_name} has arrived to meet you.\n\nPurpose: {purpose}\nPlease be informed."
    msg = MIMEText(body)
    msg["Subject"] = subject
    msg["From"] = sender
    msg["To"] = receiver

    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(sender, EMAIL_PASSWORD)
        server.sendmail(sender, [receiver], msg.as_string())
        server.quit()
    except Exception as e:
        print(f"[EMAIL ERROR]: {e}")


def get_weather_and_time():
    try:
        weather_url = f"http://api.weatherapi.com/v1/current.json?key={WEATHER_API_KEY}&q={CITY_NAME}"
        response = requests.get(weather_url)
        data = response.json()
        temp_c = data['current']['temp_c']
        condition = data['current']['condition']['text']
        local_time = data['location']['localtime']
        return f"The current temperature in {CITY_NAME} is {temp_c}°C with {condition}. Local time is {local_time}."
    except Exception as e:
        print(f"[WEATHER ERROR]: {e}")
        return "Sorry, I couldn't fetch the current weather and time."

def extract_city_name_from_query(query):
    query = query.lower()

    # List of trigger words/phrases before the city
    trigger_patterns = [
        r"weather in ([a-zA-Z\s]+)",
        r"weather of ([a-zA-Z\s]+)",
        r"temperature in ([a-zA-Z\s]+)",
        r"temperature of ([a-zA-Z\s]+)",
        r"time in ([a-zA-Z\s]+)",
        r"time of ([a-zA-Z\s]+)"
    ]

    for pattern in trigger_patterns:
        match = re.search(pattern, query)
        if match:
            return match.group(1).strip().title()

    # fallback
    return CITY_NAME



def is_stemrobo_query(query):
    query = query.lower()
    stemrobo_keywords = [
        "stemrobo", "stem robo", "stem", "company", "mission", "labs", "ai lab", 
        "NEP", "atl", "vision", "ar", "vr", "products", "services", "focus", 
        "curriculum", "learn", "students", "noida", "established", "cin", 
        "registration", "location", "tinkering", "business model", "features"
    ]

    # Only match if query explicitly refers to stemrobo or its context
    if "stemrobo" in query or "stem robo" in query:
        return True

    # Match if a Stemrobo keyword is present AND "founder", "location", etc. is also present
    for keyword in stemrobo_keywords:
        if keyword in query:
            return True
    return False


def answer_from_stemrobo_data(query):
    query = query.lower()
    if "mission" in query:
        return stemrobo_data["company_overview"]["mission"]
    if "established" in query:
        return f"STEMROBO Technologies was established on {stemrobo_data['company_overview']['established']}."
    if "location" in query:
        return f"The company is located in {stemrobo_data['company_overview']['location']}."
    return stemrobo_data["company_overview"]["description"]

def main():
    while True:
        image_path = capture_image()
        if not image_path:
            speak_async("Could not capture your image.")
            continue

        face_img, face_location, full_img = extract_face(image_path)
        if face_img is None:
            speak_async("I could not detect a face.")
            continue

        rgb_img = cv2.cvtColor(full_img, cv2.COLOR_BGR2RGB)
        encodings = face_recognition.face_encodings(rgb_img, [face_location])
        if not encodings:
            speak_async("Sorry, I couldn't process your face.")
            continue

        distances = face_recognition.face_distance(known_face_encodings, encodings[0])
        visitor_name = "Unknown"

        if len(distances) > 0:
            min_distance = min(distances)
            best_match_index = distances.tolist().index(min_distance)
            if min_distance < 0.45:
                visitor_name = known_face_names[best_match_index]
                speak_async(f"Hello {visitor_name}, welcome back to Stemrobo Technologies!")

        if visitor_name == "Unknown":
            speak_async("Hello! Welcome to Stemrobo Technologies. What is your name?")
            while True:
                visitor_name = listen()
                if visitor_name != "Unknown":
                    break
                speak_async("Sorry, I didn't catch that. Can you please repeat your name?")

            known_face_encodings.append(encodings[0])
            known_face_names.append(visitor_name)
            with open(encodings_file, "wb") as f:
                pickle.dump({"encodings": known_face_encodings, "names": known_face_names}, f)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(IMAGE_DIR, f"{visitor_name}_{timestamp}.jpg")
        top, right, bottom, left = face_location
        cv2.imwrite(filename, full_img)

        speak_async("Who would you like to meet today?")
        person_to_meet = listen()

        speak_async("What is the purpose of your visit?")
        purpose = listen()

        if person_to_meet != "Unknown":
            send_email_notification(person_to_meet, visitor_name, purpose)
            speak_async(f"Thank you. I have informed {person_to_meet} about your arrival.")

        wb = load_workbook(excel_file)
        ws = wb.active
        ws.append([visitor_name, timestamp, filename, person_to_meet, purpose])
        wb.save(excel_file)

        speak_async("How may I help you today?")

        while True:
            query = listen()
            if query != "Unknown":
                if "weather" in query.lower() or "time" in query.lower():
                    city = extract_city_name_from_query(query)
                    response = get_weather_and_time(city)

                else:
                    response = answer_from_stemrobo_data(query) if is_stemrobo_query(query) else ask_openai(query)
                speak_async(response)

                if any(phrase in query.lower() for phrase in ["bye", "no", "nothing", "thank you", "thanks", "see you"]):
                    speak_async("You're welcome! It was a pleasure talking to you. Have a great day ahead. Goodbye!")
                    break
                else:
                    speak_async("Anything else I can help you with?")
            else:
                speak_async("Sorry, I didn't catch that. Could you say it again?")

if __name__ == "__main__":
    main()
